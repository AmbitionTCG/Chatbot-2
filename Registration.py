
from openpyxl import Workbook, load_workbook

RegVar = False


def RegFunc(write="", issuecat=""):

    global RegVar

    if issuecat != "registrera för prova på dag":
        with open('User_inputs.txt', 'a') as f:
            f.write('\n' + write)
            print("file succesfully written in!")
        RegVar = False
    else:
        RegVar = True



# Function to process input and save to Excel
def process_and_save_to_excel(input_data, excel_file="output.xlsx"):
    # Split input into name, surname, and email
    try:
        name, surname, email = input_data.split()
    except ValueError:
        print("Invalid input format. Please provide 'Name Surname Email'.")
        return
    
    # Try to load an existing workbook, else create a new one
    try:
        workbook = load_workbook(excel_file)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        # Write headers if new file
        sheet.append(["Name", "Surname", "Email"])
    
    # Append the data
    sheet.append([name, surname, email])
    
    # Save the workbook
    workbook.save(excel_file)
    print(f"Data saved to {excel_file}.")

# Example usage
input_string = "John Doe john.doe@example.com"  # Example input
process_and_save_to_excel(input_string)

